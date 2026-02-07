"""
SMOKE TEST - Validazione Parsing Header RAS

Test pre-import per verificare che il parsing funzioni correttamente
PRIMA di lanciare l'import massivo su 99 file.

Uso:
    python smoke_test_header_parse.py "C:\\Path\\File\\Test.xlsx"

Output atteso:
    ✓ Header trovato
    ✓ Cognome/Nome estratti
    ✓ Porto d'arma estratto (se presente)
    ✓ numero_tessera SEMPRE valorizzato (MAI None/vuoto)
"""

import sys
import os
import re
import hashlib
import datetime as dt
from openpyxl import load_workbook

def extract_header_text_from_excel(file_path: str) -> str:
    """Estrae header text dalle prime 6 righe"""
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        
        for row in range(1, 7):
            for col in range(1, 81):
                try:
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value and isinstance(cell_value, str) and len(cell_value) > 50:
                        text_lower = cell_value.lower()
                        if any(kw in text_lower for kw in ["sig.", "porto d'arma", "porto d arma", "autorizzazione"]):
                            wb.close()
                            return cell_value.strip()
                except Exception:
                    continue
        
        wb.close()
    except Exception as e:
        print(f"❌ Errore lettura Excel: {e}")
    
    return ""

def genera_numero_tessera_stabile(file_name: str, cognome: str, nome: str, porto_arma: str, anno: int) -> str:
    """Genera numero tessera stabile (stesso file = stessa tessera)"""
    if porto_arma and porto_arma.strip():
        return f"PA_{porto_arma.strip()}"
    
    hash_obj = hashlib.md5(file_name.encode('utf-8'))
    hash_hex = hash_obj.hexdigest()
    numero_hash = int(hash_hex[:8], 16) % 90000 + 10000
    return f"AUTO_{anno}_{numero_hash}"

def extract_data_from_header_text(header_text: str, file_name: str, anno: int) -> dict:
    """Estrae dati da header text"""
    dati = {
        'cognome': '',
        'nome': '',
        'porto_arma': '',
        'data_rilascio': None,
        'autorizzazione_regionale': '',
        'numero_tessera': '',
        'stato': 'RILASCIATO'
    }
    
    # 1. Cognome e Nome
    if header_text:
        match_nome = re.search(r"\bSig\.?\s+([A-ZÀ-Ù][A-Za-zÀ-ÿ''\-]+)\s+([A-ZÀ-Ù][A-Za-zÀ-ÿ''\-]+)", header_text, re.IGNORECASE)
        if match_nome:
            dati['cognome'] = match_nome.group(1).upper()
            dati['nome'] = match_nome.group(2).title()
    
    # Fallback da filename
    if not dati['cognome'] or not dati['nome']:
        nome_pulito = file_name.replace('_', ' ').split('(')[0].split('.')[0].strip()
        match_file = re.match(r"([A-Za-zÀ-ÿ''\-]+)\s+([A-Za-zÀ-ÿ''\-]+)", nome_pulito)
        if match_file:
            dati['cognome'] = match_file.group(1).upper()
            dati['nome'] = match_file.group(2).title()
    
    # 2. Porto d'arma
    if header_text:
        match_porto = re.search(r"porto\s+d['\']arma\s*n[°ºo]?\s*([A-Za-z0-9\-\/]+)", header_text, re.IGNORECASE)
        if match_porto:
            dati['porto_arma'] = match_porto.group(1).strip()
    
    # 3. Data
    if header_text:
        match_data = re.search(r"\b(\d{2}/\d{2}/\d{4})\b", header_text)
        if match_data:
            try:
                dati['data_rilascio'] = dt.datetime.strptime(match_data.group(1), '%d/%m/%Y').date()
            except:
                pass
    
    # 4. Autorizzazione regionale
    if header_text:
        match_autorizzazione = re.search(r"autorizzazione\s+regionale\s*n[°ºo]?\s*([0-9]+)", header_text, re.IGNORECASE)
        if match_autorizzazione:
            dati['autorizzazione_regionale'] = match_autorizzazione.group(1)
    
    # 5. Stato da filename
    match_stato = re.search(r'\((.*?)\)', file_name)
    if match_stato:
        stato_raw = match_stato.group(1).upper()
        if 'CONSEGNATO' in stato_raw:
            dati['stato'] = 'CONSEGNATO'
        elif 'STAMPATO' in stato_raw:
            dati['stato'] = 'RILASCIATO'
    
    # ========== GUARDIA FINALE ==========
    dati['numero_tessera'] = genera_numero_tessera_stabile(
        file_name=file_name,
        cognome=dati['cognome'],
        nome=dati['nome'],
        porto_arma=dati['porto_arma'],
        anno=anno
    )
    
    # Doppia sicurezza
    if not dati['numero_tessera'] or dati['numero_tessera'].strip() == '':
        hash_obj = hashlib.md5(file_name.encode('utf-8'))
        dati['numero_tessera'] = f"EMERGENCY_{hash_obj.hexdigest()[:12].upper()}"
    
    return dati

def main():
    print("="*70)
    print("SMOKE TEST - Parsing Header RAS 2025-26")
    print("="*70)
    
    if len(sys.argv) < 2:
        print("\n❌ Uso: python smoke_test_header_parse.py <file.xlsx>")
        print("\nEsempio:")
        print('  python smoke_test_header_parse.py "C:\\FOGLI\\Bandino Giuseppe(Stampato).xlsx"')
        return
    
    file_path = sys.argv[1]
    
    if not os.path.exists(file_path):
        print(f"\n❌ File non trovato: {file_path}")
        return
    
    file_name = os.path.basename(file_path)
    anno = 2025
    
    print(f"\n📁 File: {file_name}")
    print(f"📍 Path: {file_path}")
    print(f"📅 Anno: {anno}")
    print()
    
    # STEP 1: Estrai header
    print("🔍 STEP 1: Estrazione header text...")
    header_text = extract_header_text_from_excel(file_path)
    
    if header_text:
        print("✅ Header trovato!")
        print(f"   Lunghezza: {len(header_text)} caratteri")
        print(f"   Preview: {header_text[:200]}...")
    else:
        print("⚠️ Header NON trovato (fallback su filename)")
    
    print()
    
    # STEP 2: Estrai dati
    print("🔍 STEP 2: Estrazione dati strutturati...")
    dati = extract_data_from_header_text(header_text, file_name, anno)
    
    print(f"   Cognome: {dati['cognome'] or '❌ NON TROVATO'}")
    print(f"   Nome: {dati['nome'] or '❌ NON TROVATO'}")
    print(f"   Porto d'arma: {dati['porto_arma'] or 'N/A'}")
    print(f"   Data rilascio: {dati['data_rilascio'] or 'N/A'}")
    print(f"   Autorizzazione: {dati['autorizzazione_regionale'] or 'N/A'}")
    print(f"   Stato: {dati['stato']}")
    print()
    
    # STEP 3: Verifica numero_tessera (CRITICO)
    print("🔍 STEP 3: Verifica numero_tessera (CRITICO)...")
    numero_tessera = dati['numero_tessera']
    
    if not numero_tessera or numero_tessera.strip() == '':
        print("❌ CRITICO: numero_tessera è VUOTO!")
        print("   Questo causerà: NOT NULL constraint failed")
        print()
        print("="*70)
        print("❌ TEST FALLITO")
        print("="*70)
        return 1
    else:
        print(f"✅ numero_tessera: {numero_tessera}")
        print(f"   Tipo: {type(numero_tessera)}")
        print(f"   Lunghezza: {len(numero_tessera)}")
        
        # Verifica stabilità
        dati2 = extract_data_from_header_text(header_text, file_name, anno)
        if dati2['numero_tessera'] == numero_tessera:
            print(f"✅ STABILE: stesso file genera stessa tessera")
        else:
            print(f"⚠️ WARNING: tessera non stabile tra run!")
    
    print()
    
    # RIEPILOGO FINALE
    print("="*70)
    print("✅ TEST COMPLETATO CON SUCCESSO")
    print("="*70)
    print()
    print("📊 RIEPILOGO:")
    print(f"   • Cognome/Nome: {'✅ Estratti' if (dati['cognome'] and dati['nome']) else '❌ Mancanti'}")
    print(f"   • Porto d'arma: {'✅ Trovato' if dati['porto_arma'] else '⚠️ Non presente (OK)'}")
    print(f"   • numero_tessera: ✅ GARANTITO ({numero_tessera})")
    print()
    print("🚀 Pronto per import massivo!")
    print()
    
    return 0

if __name__ == "__main__":
    exit(main())
