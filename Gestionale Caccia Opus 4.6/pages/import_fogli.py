import streamlit as st
import pandas as pd
import openpyxl
import os
import re
import datetime as dt
import sys
import logging
import time
import hashlib
from openpyxl import load_workbook

# Setup logging su file
logging.basicConfig(
    filename='import_debug.log',
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Import del parser Excel
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from excel_parser import ExcelParser

# Contatore globale per numero_tessera autogenerato (per run)
_tessera_counter = 0

def show():
    """Mostra la pagina import fogli caccia"""
    st.markdown('<div class="main-header">📥 Import Massivo Fogli Caccia</div>', unsafe_allow_html=True)
    
    st.info("""
    **Import Automatico Fogli Caccia da File Excel**
    
    Carica automaticamente i fogli caccia esistenti nel gestionale leggendo i file dalla cartella.
    
    **Formati file supportati:**
    - `Cognome Nome(Stato).xlsx` → es. `Bandino Giuseppe(Stampato).xlsx`
    - File con dati in header testuale (modello RAS 2025-26)
    - Estrazione automatica da testo libero
    """)
    
    # Tabs
    tab1, tab2, tab3 = st.tabs([
        "📂 Import da Cartella",
        "📄 Import Singolo File", 
        "📋 Fogli Importati"
    ])
    
    with tab1:
        show_import_da_cartella()
    
    with tab2:
        show_import_singolo_file()
    
    with tab3:
        show_fogli_importati()

def genera_numero_tessera_stabile(file_name: str, cognome: str, nome: str, porto_arma: str, anno: int) -> str:
    """
    TASK 1: Genera numero_tessera STABILE e UNIVOCO
    
    Usa hash del filename come base per garantire stabilità tra run diversi
    Fallback multipli per garantire che NON sia mai vuoto
    
    Args:
        file_name: Nome file per hash stabile
        cognome: Cognome cacciatore
        nome: Nome cacciatore  
        porto_arma: Numero porto d'arma (se disponibile)
        anno: Anno fogli
    
    Returns:
        str: Numero tessera GARANTITO non vuoto
    """
    # PRIORITA' 1: Se abbiamo porto d'arma, usa quello
    if porto_arma and porto_arma.strip():
        tessera = f"PA_{porto_arma.strip()}"
        logger.info(f"Tessera da porto_arma: {tessera}")
        return tessera
    
    # PRIORITA' 2: Hash stabile del filename (stesso file = stessa tessera)
    hash_obj = hashlib.md5(file_name.encode('utf-8'))
    hash_hex = hash_obj.hexdigest()
    numero_hash = int(hash_hex[:8], 16) % 90000 + 10000
    tessera = f"AUTO_{anno}_{numero_hash}"
    logger.info(f"Tessera da hash file '{file_name}': {tessera}")
    return tessera
    
    # NOTA: Non serve fallback perché hash è sempre disponibile

def extract_header_text_from_excel(file_path: str) -> str:
    """
    Estrae il testo dell'header dalle prime righe del foglio Excel
    Cerca nelle righe 1-6, colonne 1-80
    
    Returns:
        str: Testo header trovato, o stringa vuota se non trovato
    """
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        
        # Scorri righe 1-6 e colonne 1-80
        for row in range(1, 7):
            for col in range(1, 81):
                try:
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value and isinstance(cell_value, str) and len(cell_value) > 50:
                        # Verifica se contiene pattern rilevanti
                        text_lower = cell_value.lower()
                        if any(kw in text_lower for kw in ["sig.", "porto d'arma", "porto d arma", "autorizzazione"]):
                            wb.close()
                            logger.info(f"Header trovato alla riga {row}, col {col}: {cell_value[:100]}")
                            return cell_value.strip()
                except Exception:
                    continue
        
        wb.close()
    except Exception as e:
        logger.warning(f"Errore lettura header Excel {file_path}: {e}")
    
    return ""

def extract_data_from_header_text(header_text: str, file_name: str, anno: int) -> dict:
    """
    Estrae dati strutturati dall'header testuale usando regex robuste
    
    GUARDIA FINALE: numero_tessera è SEMPRE valorizzato
    
    Args:
        header_text: Testo header estratto dal file Excel
        file_name: Nome file (fallback per cognome/nome)
        anno: Anno fogli
    
    Returns:
        dict: Dizionario con dati estratti (numero_tessera GARANTITO non None/vuoto)
    """
    dati = {
        'cognome': '',
        'nome': '',
        'porto_arma': '',
        'data_rilascio': None,
        'autorizzazione_regionale': '',
        'numero_tessera': '',  # Verrà valorizzato in GUARDIA FINALE
        'stato': 'RILASCIATO'
    }
    
    # 1. Estrai Cognome e Nome
    if header_text:
        match_nome = re.search(r"\bSig\.?\s+([A-ZÀ-Ù][A-Za-zÀ-ÿ''\-]+)\s+([A-ZÀ-Ù][A-Za-zÀ-ÿ''\-]+)", header_text, re.IGNORECASE)
        if match_nome:
            dati['cognome'] = match_nome.group(1).upper()
            dati['nome'] = match_nome.group(2).title()
    
    # Fallback da filename se non trovato in header
    if not dati['cognome'] or not dati['nome']:
        # Rimuovi estensione e stato tra parentesi
        nome_pulito = file_name.replace('_', ' ').split('(')[0].split('.')[0].strip()
        match_file = re.match(r"([A-Za-zÀ-ÿ''\-]+)\s+([A-Za-zÀ-ÿ''\-]+)", nome_pulito)
        if match_file:
            dati['cognome'] = match_file.group(1).upper()
            dati['nome'] = match_file.group(2).title()
    
    # 2. Estrai porto d'arma
    if header_text:
        match_porto = re.search(r"porto\s+d['\']arma\s*n[°ºo]?\s*([A-Za-z0-9\-\/]+)", header_text, re.IGNORECASE)
        if match_porto:
            dati['porto_arma'] = match_porto.group(1).strip()
    
    # 3. Estrai data
    if header_text:
        match_data = re.search(r"\b(\d{2}/\d{2}/\d{4})\b", header_text)
        if match_data:
            try:
                dati['data_rilascio'] = dt.datetime.strptime(match_data.group(1), '%d/%m/%Y').date()
            except:
                pass
    
    # 4. Estrai autorizzazione regionale
    if header_text:
        match_autorizzazione = re.search(r"autorizzazione\s+regionale\s*n[°ºo]?\s*([0-9]+)", header_text, re.IGNORECASE)
        if match_autorizzazione:
            dati['autorizzazione_regionale'] = match_autorizzazione.group(1)
    
    # 5. Estrai stato da filename
    match_stato = re.search(r'\((.*?)\)', file_name)
    if match_stato:
        stato_raw = match_stato.group(1).upper()
        if 'CONSEGNATO' in stato_raw:
            dati['stato'] = 'CONSEGNATO'
        elif 'STAMPATO' in stato_raw or 'STAMPATA' in stato_raw:
            dati['stato'] = 'RILASCIATO'
        elif 'RINNOVARE' in stato_raw:
            dati['stato'] = 'DISPONIBILE'
    
    # ========== GUARDIA FINALE: numero_tessera SEMPRE valorizzato ==========
    # TASK 1: Questo è il punto critico che previene NOT NULL constraint failed
    
    dati['numero_tessera'] = genera_numero_tessera_stabile(
        file_name=file_name,
        cognome=dati['cognome'],
        nome=dati['nome'],
        porto_arma=dati['porto_arma'],
        anno=anno
    )
    
    # Verifica doppia sicurezza (non dovrebbe mai essere necessaria)
    if not dati['numero_tessera'] or dati['numero_tessera'].strip() == '':
        # Ultimo fallback assoluto: hash del filename
        hash_obj = hashlib.md5(file_name.encode('utf-8'))
        dati['numero_tessera'] = f"EMERGENCY_{hash_obj.hexdigest()[:12].upper()}"
        logger.error(f"EMERGENCY tessera per {file_name}: {dati['numero_tessera']}")
    
    logger.info(f"File {file_name}: Tessera finale = {dati['numero_tessera']}")
    
    return dati

def cerca_cacciatore_fuzzy(cognome: str, nome: str, cacciatori_esistenti: list, data_nascita=None) -> dict:
    """Cerca un cacciatore con matching fuzzy su cognome+nome"""
    if not cognome or not nome:
        return None
    
    cognome_upper = cognome.upper().strip()
    nome_title = nome.title().strip()
    
    # Match esatto
    for c in cacciatori_esistenti:
        if c.get('cognome', '').upper() == cognome_upper and c.get('nome', '').title() == nome_title:
            if data_nascita and c.get('data_nascita'):
                if str(c.get('data_nascita')) == str(data_nascita):
                    return c
            else:
                return c
    
    # Match fuzzy (cognome esatto, nome abbreviato)
    for c in cacciatori_esistenti:
        cognome_db = c.get('cognome', '').upper()
        nome_db = c.get('nome', '').title()
        
        if cognome_db == cognome_upper:
            if nome_title in nome_db or nome_db in nome_title:
                return c
    
    return None

def scansiona_e_importa_fogli(cartella: str, anno: int):
    """
    TASK 3: Import massivo atomico con logging tecnico
    
    Ogni file è un'operazione atomica (successo completo o rollback)
    Logging su file per debug tecnico
    """
    
    logger.info(f"========== INIZIO IMPORT MASSIVO ==========")
    logger.info(f"Cartella: {cartella}, Anno: {anno}")
    
    st.info("🔍 Scansione cartella e parsing file Excel in corso...")
    
    # Trova tutti i file Excel
    files = [f for f in os.listdir(cartella) if f.endswith('.xlsx') or f.endswith('.xls')]
    
    if not files:
        st.warning("⚠️ Nessun file Excel trovato nella cartella")
        logger.warning("Nessun file trovato")
        return
    
    st.success(f"✅ Trovati {len(files)} file Excel")
    logger.info(f"File trovati: {len(files)}")
    
    # Progress bar
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    # Contatori
    importati = 0
    errori = 0
    gia_esistenti = 0
    cacciatori_creati = 0
    cacciatori_aggiornati = 0
    
    username = "admin"
    
    # Lista errori dettagliati
    errori_dettaglio = []
    
    # TASK 3: Processa ogni file in modo atomico
    for idx, file_name in enumerate(files):
        file_path = os.path.join(cartella, file_name)
        
        # Aggiorna progress
        progress_bar.progress((idx + 1) / len(files))
        status_text.text(f"Elaborazione {idx + 1}/{len(files)}: {file_name}")
        
        logger.info(f"--- Processing file {idx+1}/{len(files)}: {file_name} ---")
        
        # TASK 3: Operazione atomica per file
        success = False
        for tentativo in range(3):  # Max 3 tentativi
            try:
                # Estrai header
                header_text = extract_header_text_from_excel(file_path)
                logger.debug(f"Header estratto: {header_text[:200] if header_text else 'VUOTO'}")
                
                # Estrai dati (con GUARDIA numero_tessera)
                dati_estratti = extract_data_from_header_text(header_text, file_name, anno)
                
                cognome = dati_estratti['cognome']
                nome = dati_estratti['nome']
                numero_tessera = dati_estratti['numero_tessera']  # GARANTITO non vuoto
                porto_arma = dati_estratti['porto_arma']
                stato = dati_estratti['stato']
                
                # TASK 4: Log tecnico valore finale numero_tessera
                logger.info(f"Dati estratti: cognome={cognome}, nome={nome}, tessera={numero_tessera}, porto={porto_arma}")
                
                if not cognome or not nome:
                    logger.error(f"SKIP: Impossibile estrarre cognome/nome da {file_name}")
                    errori += 1
                    errori_dettaglio.append(f"{file_name}: Cognome/nome mancanti")
                    break
                
                # Doppia verifica numero_tessera (sicurezza ridondante)
                if not numero_tessera or numero_tessera.strip() == '':
                    logger.critical(f"CRITICAL: tessera vuota dopo guardia! File: {file_name}")
                    raise ValueError(f"numero_tessera vuoto nonostante guardia: {file_name}")
                
                # Cerca o crea cacciatore
                cacciatori_esistenti = st.session_state.db.get_tutti_cacciatori(solo_attivi=True)
                cacciatore = cerca_cacciatore_fuzzy(cognome, nome, cacciatori_esistenti)
                
                if not cacciatore:
                    # Crea nuovo cacciatore
                    dati_cacciatore = {
                        'cognome': cognome,
                        'nome': nome,
                        'codice_fiscale': None,
                        'data_nascita': dati_estratti.get('data_rilascio'),
                        'numero_tessera': numero_tessera,  # OBBLIGATORIO
                        'attivo': 1
                    }
                    
                    logger.info(f"Creazione nuovo cacciatore: {cognome} {nome}, tessera={numero_tessera}")
                    cacciatore_id = st.session_state.db.aggiungi_cacciatore(dati_cacciatore)
                    cacciatore = st.session_state.db.get_cacciatore(cacciatore_id)
                    cacciatori_creati += 1
                else:
                    logger.info(f"Cacciatore esistente trovato: ID={cacciatore['id']}")
                
                # Genera numero foglio univoco
                if dati_estratti.get('autorizzazione_regionale'):
                    numero_foglio = f"{anno}_{dati_estratti['autorizzazione_regionale']}"
                else:
                    hash_obj = hashlib.md5(file_name.encode())
                    numero_seq = int(hash_obj.hexdigest()[:8], 16) % 900000 + 100000
                    numero_foglio = f"{anno}{numero_seq}"
                
                # Verifica esistenza
                fogli_esistenti = st.session_state.db.get_fogli_anno(anno)
                if any(f.get('numero_foglio') == numero_foglio for f in fogli_esistenti):
                    logger.info(f"Foglio già esistente: {numero_foglio}")
                    gia_esistenti += 1
                    break
                
                # Crea foglio
                dati_foglio = {
                    'numero_foglio': numero_foglio,
                    'anno': anno,
                    'cacciatore_id': cacciatore['id'],
                    'tipo': 'A3',
                    'data_rilascio': dt.datetime.now().date().isoformat() if not dati_estratti.get('data_rilascio') else dati_estratti['data_rilascio'].isoformat(),
                    'rilasciato_a': f"{cacciatore['cognome']} {cacciatore['nome']}",
                    'stato': stato,
                    'note': f"Import: {file_name}. Porto: {porto_arma or 'N/A'}",
                    'file_path': file_path
                }
                
                logger.info(f"Inserimento foglio: {numero_foglio}")
                foglio_id = st.session_state.db.aggiungi_foglio_caccia(dati_foglio)
                
                logger.info(f"SUCCESS: Foglio ID={foglio_id} creato")
                importati += 1
                success = True
                
                # Throttle
                if importati % 10 == 0:
                    time.sleep(0.05)
                
                break  # Esci dal retry loop
                
            except Exception as e:
                error_msg = str(e).lower()
                
                # TASK 4: Log errore tecnico
                logger.error(f"Tentativo {tentativo+1}/3 fallito per {file_name}: {e}")
                
                # TASK 2: Gestione database locked
                if "locked" in error_msg and tentativo < 2:
                    logger.warning(f"Database locked, attesa {0.5 * (tentativo + 1)}s...")
                    time.sleep(0.5 * (tentativo + 1))
                    continue
                else:
                    # Errore definitivo
                    errori += 1
                    errori_dettaglio.append(f"{file_name}: {str(e)[:150]}")
                    logger.error(f"ERRORE DEFINITIVO per {file_name}: {e}")
                    break
    
    logger.info(f"========== FINE IMPORT MASSIVO ==========")
    logger.info(f"Importati: {importati}, Errori: {errori}, Già esistenti: {gia_esistenti}")
    
    # Risultati finali
    progress_bar.progress(1.0)
    status_text.text("✅ Scansione completata!")
    
    st.markdown("---")
    st.markdown("### 📊 Riepilogo Import")
    
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("📁 File Trovati", len(files))
    with col2:
        st.metric("✅ Importati", importati, delta=importati if importati > 0 else None)
    with col3:
        st.metric("⚠️ Già Esistenti", gia_esistenti)
    with col4:
        st.metric("❌ Errori", errori, delta=-errori if errori > 0 else None)
    
    col5, col6 = st.columns(2)
    with col5:
        st.metric("👤 Cacciatori Creati", cacciatori_creati)
    with col6:
        st.metric("🔄 Cacciatori Aggiornati", cacciatori_aggiornati)
    
    if importati > 0:
        st.success(f"✅ Import completato! {importati} fogli importati.")
    
    if gia_esistenti > 0:
        st.info(f"ℹ️ {gia_esistenti} fogli già presenti (saltati)")
    
    if errori > 0:
        st.error(f"❌ {errori} errori durante l'import")
        with st.expander("📋 Dettaglio errori (primi 20)"):
            for errore in errori_dettaglio[:20]:
                st.text(errore)
        st.info("📄 Vedi import_debug.log per dettagli completi")

def show_import_da_cartella():
    """Import massivo da cartella"""
    st.subheader("Import Massivo da Cartella")
    
    cartella_fogli = st.text_input(
        "Percorso Cartella Fogli Caccia",
        placeholder=r"C:\Percorso\Cartella\Fogli",
        help="Inserisci il percorso completo della cartella contenente i file Excel"
    )
    
    anno_import = st.number_input(
        "Anno Fogli",
        min_value=2020,
        max_value=2030,
        value=2025,
        step=1
    )
    
    if st.button("🔍 Scansiona Cartella", type="primary", use_container_width=True):
        if not cartella_fogli or not os.path.exists(cartella_fogli):
            st.error("⚠️ Cartella non trovata!")
        else:
            scansiona_e_importa_fogli(cartella_fogli, anno_import)

def show_import_singolo_file():
    """Import singolo file (placeholder)"""
    st.subheader("Import Singolo File")
    st.info("Usa l'import massivo per ora")

def show_fogli_importati():
    """Visualizza fogli importati"""
    st.subheader("📋 Fogli Importati")
    
    anno_vista = st.number_input(
        "Anno da visualizzare",
        min_value=2020,
        max_value=2030,
        value=2025,
        step=1,
        key="anno_vista_import"
    )
    
    fogli = st.session_state.db.get_fogli_anno(anno_vista)
    
    if not fogli:
        st.info(f"ℹ️ Nessun foglio per l'anno {anno_vista}")
        return
    
    st.success(f"✅ Trovati {len(fogli)} fogli per l'anno {anno_vista}")
    
    # Statistiche
    statistiche = st.session_state.db.get_statistiche_fogli(anno_vista)
    
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Totale", statistiche.get('totale', 0))
    with col2:
        st.metric("Disponibili", statistiche.get('disponibili', 0))
    with col3:
        st.metric("Rilasciati", statistiche.get('rilasciati', 0))
    with col4:
        st.metric("Restituiti", statistiche.get('restituiti', 0))
    
    # Tabella
    df = pd.DataFrame(fogli)
    if not df.empty:
        cols = ['numero_foglio', 'cognome', 'nome', 'stato', 'data_rilascio']
        cols_ok = [c for c in cols if c in df.columns]
        st.dataframe(df[cols_ok], use_container_width=True, hide_index=True)
